from gemini_api.authentication import Authentication
from gemini_api.endpoints.order import Order
from openpyxl import load_workbook
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
import sys

coins = []
principle = 0

def parse_file(file_path):
    global principle
    #Load the spreadsheet and make it active.
    wb = load_workbook(filename = str(file_path).replace("'",'').replace("]", '').replace("[", ''))
    sheet = wb.active

    #This loops goes over each column and check if the first row has the word Balance in it. If so, it stores it in a list.
    #Balance has the total amount of a given coin in the bottom-most row.
    for column in sheet.iter_cols(min_row=1, min_col=1, max_row=sheet.max_row, max_col=sheet.max_column):
        col_value = str(column[0].value)  #This converts the cell's text into usable strings.
        if "Balance" in col_value:
            coins.append([col_value])
            coins[-1].append(str(column[-1].value))  #This grabs the total current balance of a coin/currency
        if "CAD Amount CAD" in col_value:
            for item in column:
                if "CAD Amount" not in str(item.value) and not isinstance(item.value, type(None)) and float(item.value) <= 0:
                    principle += float(item.value)
        principle = round(principle, 2)
        principle = abs(principle)
    wb.close()
    principle = round(principle, 3)


#This class is for the root window.
class Window(QWidget):
    def __init__(self, parent = None):
        super(Window, self).__init__(parent)  #This loads the __init__ from the QWidget class.
        self.setWindowTitle('Gemini ROI Tool')
        self.resize(800, 800)
        self.list_widget = QListWidget()
        grid = QGridLayout()

        #Create the needed widgets.
        self.file_btn = QPushButton("Load File")  #The button object and it's display text.

        #Add any functionality to widgets
        self.file_btn.clicked.connect(self.load_file)  #For when the button is clicked, it executes the passed function.

        #Add the widgets to the grid.
        grid.addWidget(self.file_btn)  #Adds the button to the layout.
        grid.addWidget(self.list_widget, 1, 0, 5, 5)

        self.setLayout(grid)  #This guy loads the layout and it's widgets.

    #This method is for when the relevant button is clicked. It allows the user to select the file they want to parse and then parses it.
    def load_file(self):
        global principle
        dlg = QFileDialog(self)
        dlg.setFileMode(QFileDialog.AnyFile)
        dlg.setNameFilter("transaction_history (*.xlsx)")
        if dlg.exec_():
            file_path = dlg.selectedFiles()
            parse_file(file_path)
            # self.list_widget.addItems(str(x) for item in coins for x in item)
            for item in coins:
                data = item[0:2]
                data = str(data).replace("[", '').replace("]", '').replace("'", '')
                self.list_widget.addItem(data)
        principle = "$$ Principle, " + str(principle)
        self.list_widget.addItem(principle)


def main():
    app = QApplication(sys.argv)
    root = Window()
    root.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()
